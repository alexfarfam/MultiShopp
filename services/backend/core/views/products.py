from typing import Dict
from json import loads
from datetime import datetime
import base64
from typing import Dict
import os

from django.utils.timezone import now
from django.conf import settings
from django.db.models.manager import BaseManager
from django.db.models import F, Avg, Sum, Q, IntegerField, Count
from django.db.models.functions import Coalesce
from django.http import HttpRequest, FileResponse
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.decorators import permission_classes
from django.core.files.base import ContentFile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .permissions.custom import IsProviderUser
from ..models.orders import OrderDetail
from ..models.categories import SubCategory
from ..models.products import Product, ProductImage, ProductOffer, ProductOption
from ..models.auth import History, CustomUser, Action

"""
Crud API for the product module.
"""

def get_product(id:int, request:HttpRequest):
    product:Product=Product.objects.filter(pk=id).values("id", "title", "header_tag", "description", "price", "reference_price", "show_offerts","discount", "main_image", "subcategory__id", "subcategory__title")[0]
    product["main_image"]=f"{request.scheme}://{request.get_host()}/{product['main_image'] if product['main_image'] else ''}"

    return product

def generate_excel(products:BaseManager[Product]):
    # Crea un nuevo libro y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Define los encabezados de las columnas
    headers = [
        "Titulo", "Tag", "Precio", "Precio Referencia", "Descuento",
        "SubCategoria", "Fecha Registro", "Fecha Actualizacion", "Responsable"
    ]

    # Aplica estilos a los encabezados
    # Define el título del archivo
    title = "Reporte de Productos"
    title_font = Font(size=17, color="FFFFFF", bold=True)
    title_fill = PatternFill(start_color="544fbd", end_color="544fbd", fill_type="solid")
    title_alignment = Alignment(horizontal="center")

    # Inserta el título en la primera fila
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Escribe los encabezados en la primera fila
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajusta el ancho de la columna

    # Escribe los datos de los productos en las filas siguientes
    for row_num, product in enumerate(products, 4):
        ws.cell(row=row_num, column=1, value=product.title)
        ws.cell(row=row_num, column=2, value=product.header_tag)
        ws.cell(row=row_num, column=3, value="${:0,.2f}".format(product.price))
        ws.cell(row=row_num, column=4, value="${:0,.2f}".format(product.reference_price))
        ws.cell(row=row_num, column=5, value=product.discount)
        ws.cell(row=row_num, column=6, value=product.subcategory.title)
        ws.cell(row=row_num, column=10, value=product.creation_date.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=11, value=product.last_update_date.strftime("%Y-%m-%d %H:%M:%S") if product.last_update_date else "")
        ws.cell(row=row_num, column=12, value=product.responsible.username)

    # Define el path del archivo
    fname=f"productos{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', fname)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Guarda el archivo Excel
    wb.save(file_path)
        
    return file_path

class Products_APIView(APIView):
    def get(self, request:HttpRequest, action, *args, **kwargs):
        try:
            responsible:CustomUser=request.user
            product_id=request.GET.get("product_id")
            if action == "products":
                obj:BaseManager[Product]=Product.objects.filter(pk=product_id, responsible=responsible) if product_id else Product.objects.filter(responsible=responsible)

                data=obj.values("id", "title", "header_tag", "description", "price", "reference_price", "show_offerts","discount", "main_image", "subcategory__id", "subcategory__title") 
                data=list(data)
                data_len=len(data)
                for x in range(data_len):
                    data[x]["main_image"]=f"{request.scheme}://{request.get_host()}/{data[x]['main_image'] if data[x]['main_image'] else ''}"

                if product_id:
                    if data_len == 0:
                        return Response(f"Producto con ID '{product_id}' no encontrado!", status=status.HTTP_404_NOT_FOUND)
                    data=data[0]
                
                return Response(data, status=status.HTTP_200_OK)
            elif action == "images":
                data:BaseManager[ProductImage]=ProductImage.objects.filter(product_id=product_id).values("image", "id")
                data=list(data)
                data_len=len(data)
                for x in range(data_len):
                    data[x]["name"]=f"{data[x]['image'].split('/')[1] if data[x]['image'] else ''}"
                    data[x]["image"]=f"{request.scheme}://{request.get_host()}/{data[x]['image'] if data[x]['image'] else ''}"

                return Response(data, status=status.HTTP_200_OK)
    
            elif action == "public":
                filter=request.GET.get('category_id')
                filter2=request.GET.get('subcategory_id')
                filter3=request.GET.get('top')

                obj:BaseManager[Product] = None
                if filter:
                    obj=Product.objects.filter(subcategory__category_id=filter)
                elif filter2:
                    obj=Product.objects.filter(subcategory_id=filter2)
                elif filter3:
                    if filter3 == 'best-sellers':
                        top_product_ids = (
                            OrderDetail.objects
                            .values('product')
                            .annotate(total_sold=Sum('amount'))
                            .order_by('-total_sold')
                            .values_list('product', flat=True)[:10]
                        )
                        obj = Product.objects.filter(id__in=top_product_ids)

                    elif filter3 == 'discounts':
                        obj=Product.objects.filter(discount__gt=0).order_by('-creation_date')
                    elif filter3 == 'newcomers':
                        obj=Product.objects.filter(creation_date__date = now()).order_by('-creation_date')
                    elif filter3 == 'promotions':
                        obj=Product.objects.exclude(header_tag="").order_by('-creation_date')
                else:
                    obj=Product.objects.order_by('-creation_date')

                data=obj.annotate(
                    orders=Coalesce(Sum('orderdetail__amount', output_field=IntegerField()), 0),
                    qualification=Coalesce(Avg('comment__starts', filter=Q(comment__deleted=False), output_field=IntegerField()), 0),
                    comments_count =Coalesce(Count('comment__starts', filter=Q(comment__deleted=False), output_field=IntegerField()), 0),
                ).values(
                    "id", "orders", "qualification", "comments_count", "title", "header_tag", "price", "reference_price", "discount", "main_image", "subcategory__id", "subcategory__title", "subcategory__category_id", "subcategory__category__title", "last_update_date", "responsible__id"
                ).order_by(*('-qualification', '-orders'))[:30]
                
                data=list(data)
                data_len=len(data)
                for x in range(data_len):
                    data[x]["main_image"]=f"{request.scheme}://{request.get_host()}/{data[x]['main_image'] if data[x]['main_image'] else ''}"
                return Response(data, status=status.HTTP_200_OK)
            
            elif action == "export":
                # SE PUEDEN AGREGAR FILTROS MAS ADELANTE
                products=Product.objects.filter(responsible=responsible) 
                file_path = generate_excel(products)  # Genera el archivo Excel y obtiene la ruta

                if os.path.exists(file_path):
                    filename=os.path.basename(file_path)
                    response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
                    response['Content-Disposition'] = f'attachment; filename={filename}'
                    return response
                else:
                    return Response("File not found", status=status.HTTP_404_NOT_FOUND)
    
            elif action == "info":
                obj:BaseManager[Product]=Product.objects.filter(pk=product_id)
                if obj.count() == 0:
                    return Response(f"Producto con ID '{product_id}' no encontrado!", status=status.HTTP_404_NOT_FOUND)
                
                data=obj.annotate(
                    orders=Coalesce(Sum('orderdetail__amount', output_field=IntegerField()), 0),
                    qualification=Coalesce(Avg('comment__starts', filter=Q(comment__deleted=False), output_field=IntegerField()), 0),
                    comments_count =Coalesce(Count('comment__starts', filter=Q(comment__deleted=False), output_field=IntegerField()), 0),
                ).values("id", "orders", "qualification", "comments_count", "title","description", "header_tag", "price", "reference_price", "discount", "responsible", "main_image", "subcategory__id", "subcategory__title", "last_update_date", category__id=F("subcategory__category_id"), category__title=F("subcategory__category__title"))[0]
                
                data["main_image"]=f"{request.scheme}://{request.get_host()}/{data['main_image'] if data['main_image'] else ''}"
                
                return Response(data, status=status.HTTP_200_OK)
            
            elif action == "offers":
                data:BaseManager[ProductOffer]=ProductOffer.objects.filter(product_id=product_id).values("id", "title", "reference_price", "price", "discount", "header_tag", "amount", "product")
                return Response(data, status=status.HTTP_200_OK)
            elif action == "options":
                data:BaseManager[ProductOption]=ProductOption.objects.filter(product_id=product_id).values("id", "title", "values")
                return Response(data, status=status.HTTP_200_OK)
            elif action == "public-offers":
                data:BaseManager[ProductOffer]=ProductOffer.objects.filter(product_id=product_id).values("id", "title", "reference_price", "price", "discount", "header_tag", "amount", "product", image=F("product__main_image"))
                data=list(data)
                data_len=len(data)
                for x in range(data_len):
                    data[x]["image"]=f"{request.scheme}://{request.get_host()}/{data[x]['image'] if data[x]['image'] else ''}"

                return Response(data, status=status.HTTP_200_OK)
            
            else:
                return Response("Accion no encontrada.", status=status.HTTP_404_NOT_FOUND)
        except (ValueError, TypeError) as e:
            return Response("No encontrado!", status=status.HTTP_404_NOT_FOUND)
   
    @permission_classes([IsProviderUser])
    def post(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            responsible:CustomUser=request.user
            new_data=request.data

            new_data["subcategory"]=SubCategory.objects.get(pk=new_data["subcategory"])
            new_data["creation_date"]=now()
            new_data["last_update_date"]=now()
            new_data["responsible"]=responsible
            
            # Procesa y guarda la imagen base64 en ImageField
            image_data = new_data.pop('main_image', None)  # Extrae la imagen base64 del diccionario de datos
            images_data = new_data.pop('images', None) 
            offers_data=new_data.pop('offers', None)
            charts_data=new_data.pop('characteristics', None)
            product=Product(**new_data)
            if image_data:
                format, imgstr = image_data.split(';base64,')
                ext = format.split('/')[-1]  # Determina la extensión del archivo
                image_data = ContentFile(base64.b64decode(imgstr), name=f'product_image__{now().strftime("%Y%m%d_%H%M%S")}.{ext}')
                product.main_image.save(image_data.name, image_data, save=True)  # Guarda la imagen en el campo ImageField
            product.save()
            
            if images_data:
                pro=Product.objects.get(pk=product.id)
                for img in images_data:
                    format, imgstr = img.split(';base64,')
                    ext = format.split('/')[-1] 
                    img = ContentFile(base64.b64decode(imgstr), name=f'product_image__{now().strftime("%Y%m%d_%H%M%S")}.{ext}')
                    pi=ProductImage()
                    pi.product=pro
                    pi.image.save(img.name, img, save=True)  # Guarda la imagen en el campo ImageField
                    pi.save()
            
            if offers_data and product.show_offerts:
                for index in range(len(offers_data)):
                    offer=offers_data[index]

                    _id=offer.pop("id", None)
                    po=ProductOffer.objects.filter(pk=_id)
                    if po.first():
                        po.update(**offer)
                    else:
                        offer["product"]=Product.objects.get(pk=product.id)
                        po2=ProductOffer(**offer)
                        po2.save()     
            else:
                ProductOffer.objects.filter(product_id=product.id).delete()                 

            if charts_data:
                for index in range(len(charts_data)):
                    chart=charts_data[index]

                    _id=chart.pop("id", None)
                    po=ProductOption.objects.filter(pk=_id)
                    if po.first():
                        po.update(**chart)
                    else:
                        chart["product"]=Product.objects.get(pk=product.id)
                        po2=ProductOption(**chart)
                        po2.save()  
            history=History()
            history.module="product"
            history.action=Action.CREACION.value
            history.datetime=now()
            history.responsible=responsible
            history.save()

            resp={
                "msg": f"Producto '{product.title}' creado correctamente.",
                "data": get_product(product.id, request)
            }
            return Response(resp, status=status.HTTP_200_OK)
        except Exception as e:
            err=f"Error interno, por favor, intentelo mas tarde."
            if settings.DEBUG:
                raise e
            return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @permission_classes([IsProviderUser])
    def put(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            responsible:CustomUser=request.user
            product_id=request.GET.get('product_id')
            new_data:Dict[str,str]=request.data
            new_data["subcategory"]=SubCategory.objects.get(pk=new_data["subcategory"])
            new_data["last_update_date"]=now()
            new_data["responsible"]=responsible

            product=Product.objects.filter(pk=product_id)
            has_product=product.first()
            if has_product is None:
                return Response(f"Producto con ID '{product_id}' no existe.", status=status.HTTP_404_NOT_FOUND)
            else:
                # Procesa y guarda la imagen base64 en ImageField
                image_data = new_data.pop('main_image', None)  # Extrae la imagen base64 del diccionario de datos
                images_data = new_data.pop('images', None)
                offers_data=new_data.pop('offers', None)
                charts_data=new_data.pop('characteristics', None)

                product = Product.objects.get(pk=product_id)
                for key, value in new_data.items():
                    setattr(product, key, value)
                
                if image_data and not image_data.startswith('http'):
                    format, imgstr = image_data.split(';base64,')
                    ext = format.split('/')[-1]  # Determina la extensión del archivo
                    image_data = ContentFile(base64.b64decode(imgstr), name=f'product_image__{now().strftime("%Y%m%d_%H%M%S")}.{ext}')
          
                    product.main_image.save(image_data.name, image_data, save=True)  # Guarda la imagen en el campo ImageField
                product.save()

                if images_data: 
                    for img in images_data:
                        if not img.startswith('http'):
                            format, imgstr = img.split(';base64,')
                            ext = format.split('/')[-1]  # Determina la extensión del archivo
                            img = ContentFile(base64.b64decode(imgstr), name=f'product_image__{now().strftime("%Y%m%d_%H%M%S")}.{ext}')
                
                            pi=ProductImage()
                            pi.product=product
                            pi.image.save(img.name, img, save=True)  # Guarda la imagen en el campo ImageField
                            pi.save()
                
                if offers_data and product.show_offerts:
                    for index in range(len(offers_data)):
                        offer=offers_data[index]

                        _id=offer.pop("id", None)
                        po=ProductOffer.objects.filter(pk=_id)
                        if po.first():
                            po.update(**offer)
                        else:
                            offer["product"]=Product.objects.get(pk=product.id)
                            po2=ProductOffer(**offer)
                            po2.save()     
                else:
                    ProductOffer.objects.filter(product_id=product.id).delete()                       

                if charts_data:
                    for index in range(len(charts_data)):
                        chart=charts_data[index]

                        _id=chart.pop("id", None)
                        po=ProductOption.objects.filter(pk=_id)
                        if po.first():
                            po.update(**chart)
                        else:
                            chart["product"]=Product.objects.get(pk=product.id)
                            po2=ProductOption(**chart)
                            po2.save()  

                history=History()
                history.module="product"
                history.action=Action.ACTUALIZACION.value
                history.datetime=now()
                history.responsible=responsible
                history.save()
            
            resp = {
                "msg": f"Producto '{new_data['title']}' actualizado correctamente.",
                "data": get_product(product.id, request)
            }
            return Response(resp, status=status.HTTP_200_OK)
        except Exception as e:
            err=f"Error interno, por favor, intentelo mas tarde."
            if settings.DEBUG:
                raise e
            return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @permission_classes([IsProviderUser])
    def delete(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            responsible:CustomUser=request.user
            if action == "image":
                img_id=request.GET.get('img_id')
                img_product=ProductImage.objects.filter(pk=img_id)
                has_img=img_product.first()
                if has_img is None:
                    return Response(f"Imagen con ID '{img_id}' no existe.", status=status.HTTP_404_NOT_FOUND)
                else:
                    has_img.delete()
                    return Response(f"Imagen con ID '{img_id}' eliminada correctamente!")
            
            elif action == "offer":
                offer_id=request.GET.get('offer_id')
                po=ProductOffer.objects.filter(pk=offer_id).first()
                if po:
                    po.delete()
                return Response(f"Oferta con ID '{offer_id}' eliminada correctamente!")

            elif action == "chart":
                chart_id=request.GET.get('chart_id')
                po=ProductOption.objects.filter(pk=chart_id).first()
                if po:
                    po.delete()
                return Response(f"Caracteristica con ID '{chart_id}' eliminada correctamente!")
                       
            product_ids=loads(request.GET.get('product_ids'))
            product_titles=[]
            for product_id in product_ids:
                product=Product.objects.filter(pk=int(product_id))
                has_product=product.first()
                if has_product is None:
                    return Response(f"Producto con ID '{product_id}' no existe.", status=status.HTTP_404_NOT_FOUND)
                else:
                    product.delete()
                    history=History()
                    history.module="product"
                    history.action=Action.ELIMINACION.value
                    history.datetime=now()
                    history.responsible=responsible
                    history.save()
                product_titles.append(has_product.title)
                
            string_products_titles=', '.join(product_titles).rstrip(', ')
            return Response(f"Producto(s) '{string_products_titles}' eliminado(s) correctamente.", status=status.HTTP_200_OK)

        except Exception as e:
            if settings.DEBUG:
                raise e
            return Response(f"Error interno, por favor, intentelo mas tarde.", status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    