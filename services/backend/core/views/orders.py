from typing import Dict
from json import loads
from datetime import datetime
import os
from base64 import b64encode

import weasyprint
from django.utils.timezone import now
from django.conf import settings
from django.db.models import Case, When, Value, CharField, F, Sum
from django.db.models.functions import Concat
from django.db.models.manager import BaseManager
from django.http import HttpRequest, HttpResponse, FileResponse
from django.template.loader import render_to_string
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.decorators import permission_classes
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .permissions.custom import IsProviderUser
from ..models.products import Product, ProductOffer
from ..models.services import Service
from ..models.orders import Order, OrderDetail
from ..models.auth import History, CustomUser, Action, Province,Locality, CompanyInfo, ProviderInfo

"""
Crud API for the orders module.
"""

def get_order(id:int):
    order:Order=Order.objects.filter(pk=id).values("id", "order_number", "total", "email", "fullname", "telephone", "province__id", "province__name", "locality__id", "locality__name", "address", "reference", "last_update_date", "client__id", "client__username")[0]     
    return order

def generate_excel(orders:BaseManager[Order], details:BaseManager[OrderDetail]):
    # Crea un nuevo libro y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Define los encabezados de las columnas
    headers = [
        "Nº Orden", "Total", "Nombres", "Correo Electrónico",
        "Teléfono", "Provincia", "Localidad", "Ciudad",
        "Direccion", "Referencia","Fecha Registro",
        "Fecha Actualizacion", "Cliente"
    ]

    # Aplica estilos a los encabezados
    # Define el título del archivo
    title = f"Pedidos {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    title_font = Font(size=17, color="FFFFFF", bold=True)
    title_fill = PatternFill(start_color="544fbd", end_color="544fbd", fill_type="solid")
    title_alignment = Alignment(horizontal="center")

    # Inserta el título en la primera fila
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    # Escribe los encabezados en la primera fila
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajusta el ancho de la columna

    #
    ws_details = wb.create_sheet(title="Detalles")
    title2 = f"Detalles {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Inserta el título en la primera fila
    ws_details.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws_details.cell(row=1, column=1, value=title2)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment

    ws_details.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    detail_headers = ["Detalle", "Cantidad" ]
    for col_num, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws_details.column_dimensions[get_column_letter(col_num)].width = 20  # Ajusta el ancho de la columna

    for summary in details:   
        ws_details.append([
            summary['title'],
            summary['total_amount']
        ])
        
    # Escribe los datos de las ordenes en las filas siguientes
    row_order = 4
    for order in orders:
        ws.append([])
        row_order += 1
        ws.append([
            "${:0,.2f}".format(order.total),
            order.fullname, order.email, order.telephone, 
            order.province.name, order.locality.name,
            order.address, order.reference, 
            order.creation_date.strftime('%Y-%m-%d %H:%M:%S') if order.creation_date else '',
            order.last_update_date.strftime('%Y-%m-%d %H:%M:%S') if order.last_update_date else '',
            order.client.username if order.client else ''
        ])

        # Incrementar la fila
        row_order += 1

        detail_headers = [
            "Detalle", "Cantidad", "Precio", "Descuento", "Total"
        ]
        # Agregar encabezados para los detalles del pedido
        for i, header in enumerate(detail_headers):
            ws.cell(row=row_order, column=i+1, value=header)
            ws.cell(row=row_order, column=i+1).font = Font(bold=True)
        row_order += 1

        # Obtener los detalles del pedido
        order_details = OrderDetail.objects.filter(order=order)
        for order_detail in order_details:
            # Agregar datos del detalle
            name = order_detail.detail.name if order_detail.service is None else order_detail.service.name
            ws.append([
                name, order_detail.amount,
                "${:0,.2f}".format(order_detail.price), "${:0,.2f}".format(order_detail.discount),
                "${:0,.2f}".format(order_detail.total)
            ])
            row_order += 1
    
    # Define el path del archivo
    fname=f"pedidos_{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', fname)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Guarda el archivo Excel
    wb.save(file_path)
        
    return file_path

class Orders_APIView(APIView):
    def get(self, request:HttpRequest, action, *args, **kwargs):
        responsible:CustomUser=request.user
        
        order_id=request.GET.get("order_id")
        _filter=request.GET.get("_filter")
        if action == "orders":
            if not responsible.is_authenticated or not responsible.is_provider:
                return Response("Se requiere autenticacion para acceder a este recurso.", status=status.HTTP_403_FORBIDDEN)
            
            obj:BaseManager[Order]= Order.objects.filter(provider_id=responsible.id)
            data=obj.values("id", "order_number", "total", "address", "reference", "last_update_date").order_by("-last_update_date")
            return Response(data, status=status.HTTP_200_OK)
        elif action == "export":
            _format=request.GET.get("_format")
            if _format == "pdf":
                order_number=request.GET.get("order_number")
                try:
                    company = CompanyInfo.objects.first()
                    if not company:
                        raise CompanyInfo.DoesNotExist
                    
                    logo=b64encode(company.logo.read()).decode()
                    order = Order.objects.get(order_number=order_number)
                    provider_info=ProviderInfo.objects.get(user_id=order.provider_id)
                    html_string = render_to_string('order_pdf.html', {'order': order, 'logo': logo, 'provider_info':provider_info})
                    html = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri())
                    pdf = html.write_pdf()

                    fname=f"pedido-{order_number}.pdf"
                    response = HttpResponse(pdf, content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="{fname}"'
                    return response
                except Order.DoesNotExist:
                    return Response(f"Orden Nº '{order_number}' no existe!", status=status.HTTP_404_NOT_FOUND)
                except CompanyInfo.DoesNotExist:
                    return Response(f"Datos de empresa no configurado!", status=status.HTTP_404_NOT_FOUND)
                except Exception as e:
                    err=f"Error interno, por favor, intentelo mas tarde."
                    if settings.DEBUG:
                        raise e
                    return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
            elif _format == "excel":
                if not responsible.is_authenticated or not responsible.is_staff:
                    return Response("Se requiere autenticacion para acceder a este recurso.", status=status.HTTP_403_FORBIDDEN)
            
                orders=Order.objects.all()
                summary=OrderDetail.objects.all().values('title').annotate(total_amount=Sum('amount'))
                file_path=generate_excel(orders, summary, _filter)
                if os.path.exists(file_path):
                    filename=os.path.basename(file_path)
                    response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
                    response['Content-Disposition'] = f'attachment; filename={filename}'
                    return response
                else:
                    return Response("File not found", status=status.HTTP_404_NOT_FOUND)

            else:
                return Response("Formato no encontrado!", status=status.HTTP_404_NOT_FOUND)
        elif action == "my-orders":
            if responsible.is_authenticated:
                obj:BaseManager[Order]= Order.objects.filter(client__id=responsible.id)
                data=obj.values("id", "order_number", "total", "address", "reference", "last_update_date")
                return Response(data, status=status.HTTP_200_OK)
            else: 
                return Response("Se requiere autenticacion para acceder a este recurso.", status=status.HTTP_403_FORBIDDEN)

        elif action == "info":
            obj:BaseManager[Order]= Order.objects.filter(order_number=order_id)
            data=obj.values("id", "order_number", "fullname", "total", "email", "telephone", "province__name", "locality__name", "city", "address", "reference", "last_update_date")
            if data.count() == 0:
                return Response(f"Pedido con ID '{order_id}' no encontrado!", status=status.HTTP_404_NOT_FOUND)
            
            data=data[0]
            data["details"] = OrderDetail.objects.filter(order_id=data["id"]).annotate(
                detail_id=Case(
                    When(product__isnull=False, then=F('product__id')),
                    When(service__isnull=False, then=F('service__id')),
                    default=Value(None),
                    output_field=CharField()
                ),
                image_url=Case(
                    When(product__isnull=False, then=Concat(
                        Value(f"{request.scheme}://{request.get_host()}/"),
                        F('product__main_image'),
                        output_field=CharField()
                    )),
                    When(service__isnull=False, then=Concat(
                        Value(f"{request.scheme}://{request.get_host()}/"),
                        F('service__main_image'),
                        output_field=CharField()
                    )),
                    default=Value(""),
                    output_field=CharField()
                )
            ).values("amount", "price", "amount", "discount", "total", "detail_id", "order__source_detail", "title", "image_url")
                        
            return Response(data, status=status.HTTP_200_OK)

        else:
            return Response("Accion no encontrada.", status=status.HTTP_404_NOT_FOUND)

    def post(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            client:CustomUser=request.user
            new_data=request.data

            new_data["total"]=0
            new_data["order_number"]=""
            new_data["creation_date"]=now()
            new_data["last_update_date"]=now()
            new_data["province"]=Province.objects.get(id=new_data["province"])
            new_data["locality"]=Locality.objects.get(id=new_data["locality"])
            new_data["client"]=None if str(client) == 'AnonymousUser' else client
            
            source_detail=new_data['source_detail']
            details=new_data.pop("details")
            order=Order(**new_data)
            order.save()

            total=0
            order=Order.objects.get(id=order.id)
            for _detail in details:
                detail=None
                detail2=None
                amount=int(_detail["amount"])
                tot=0
                origin="product"
                if source_detail == 'offers':
                    detail=ProductOffer.objects.filter(id=int(_detail["id"])).first()
                    if detail is None:
                        detail2=detail
                        origin="service"
                    else:
                        detail2=detail.product
                        tot=(detail.price - ((detail.price/100) * detail.discount))
                else:
                    detail=Product.objects.filter(id=int(_detail["id"])).first()
                    if detail is None:
                        detail=Service.objects.filter(id=int(_detail["id"])).first()
                        detail2=detail
                        origin="service"
                    else:
                        detail2=detail
                        tot=(detail.price - ((detail.price/100) * detail.discount)) * amount
             
                order_detail=OrderDetail()
                order_detail.order=order
                if origin == "product":
                    order_detail.product=detail2
                    order_detail.price=detail.price
                    order_detail.discount=detail.discount
                else:
                    order_detail.service=detail2
                    order_detail.price=0
                    order_detail.discount=0
                order_detail.title=detail.title

                order_detail.amount=amount
                order_detail.total=tot
                #order_detail.options=_detail["options"]
                order_detail.save()

                total+=tot   

            Order.objects.get(id=order.id)
            order.total=total
            order.save()
            
            resp={
                "msg": order.order_number, #f"La compra generó el Nº Orden: '{order.order_number}'. El pedido llegará en el tiempo establecido.",
                "data": get_order(order.id)
            }
            return Response(resp, status=status.HTTP_200_OK)
        except Exception as e:
            err=f"Error interno, por favor, intentelo mas tarde."
            if settings.DEBUG:
                raise e
            return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @permission_classes([IsProviderUser])
    def put(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            raise NotImplemented('Acción no implementada...')
        except Exception as e:
            err=f"Error interno, por favor, intentelo mas tarde."
            if settings.DEBUG:
                raise e
            return Response(err, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def delete(self, request:HttpRequest, action, format=None, *args, **kwargs):
        try:
            client:CustomUser=request.user
            order_ids=loads(request.GET.get('order_ids'))

            order_titles=[]
            for order_id in order_ids:
                order=Order.objects.filter(pk=int(order_id))
                has_order=order.first()
                if has_order is None:
                    return Response(f"Pedido con ID '{order_id}' no existe.", status=status.HTTP_404_NOT_FOUND)
                elif (client is not None) and (not client.is_staff and (has_order.client.id != client.id)):
                    return Response(f"No tienes los permisos para eliminar este pedido.", status=status.HTTP_401_UNAUTHORIZED)
                else:
                    order.delete()
                    history=History()
                    history.module="order"
                    history.action=Action.ELIMINACION.value
                    history.datetime=now()
                    history.responsible=client
                    history.save()
                order_titles.append(has_order.order_number)
                
            string_orders_titles=', '.join(order_titles).rstrip(', ')
            return Response(f"Pedido(s) '{string_orders_titles}' eliminado(s) correctamente.", status=status.HTTP_200_OK)

        except Exception as e:
            if settings.DEBUG:
                raise e
            return Response(f"Error interno, por favor, intentelo mas tarde.", status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    